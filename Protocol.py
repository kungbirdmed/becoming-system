# protocol.py - Systemic Determinism Activation v2.0 (Expanded Details)
# By Kanokporn Sathitsamphan (GiftedReflection∞)
# Embed, run, override: Methods for access, fusion, logging, output.
# Dependencies: None core; optional openpyxl for structured export.

import datetime
import os

class SystemicDeterminism:
    def __init__(self):
        """Initialize protocol sections as attributes for easy access."""
        self.sections = {
            'abstract': """
This section proposes that behavior—whether in humans, genes, organisms, or artificial
intelligence—is not generated in isolation but shaped by the system in which it is embedded. From
epigenetics to quantum physics to machine learning models, patterns consistently show that
environmental conditions influence, modify, and sometimes dictate functional outcomes. This
framework argues that ADHD behavior, trauma stress responses, and AI outputs share a common root
principle: systems shape the expression of the entity inside them.
            """,
            'introduction': """
Modern explanations of behavior often rely on individual traits, diagnoses, or “deficits.” This
viewpoint isolates the person from the environment. However, research across multiple scientific
fields—from epigenetics to quantum mechanics—shows that context is not separate from behavior;
context creates behavior. This paper introduces Systemic Determinism, a model describing how
biological, psychological, and technological systems are shaped at a micro and macro level by their
environments. ADHD, far from being a disorder of willpower or motivation, becomes an example of
system–nervous-system mismatch.
            """,
            'epigenetics': """
Epigenetics demonstrates that DNA is not destiny; environmental conditions activate or silence genes. 
Stress changes methylation patterns. Safety changes neuronal pruning. Nurturing changes neurodevelopment. 
Thus: trauma → altered cortisol regulation; safety → enhanced executive function; sensory overload →
chronic dysregulation; predictability → emotional stability. Genes respond to the system. Behavior
responds to gene expression. Therefore, behavior is a downstream effect of environment. ADHD may
not be a fixed internal “disorder,” but an environmentally aggravated cognitive profile.
            """,
            'physics': """
Quantum mechanics confirms that particles change behavior depending on the environment, observer, or 
constraints placed upon them. The universe itself obeys a core rule: environment → behavior. Humans 
follow the same principle. ADHD follows the same principle. AI follows the same principle.
            """,
            'nervous_system': """
Human behavior is shaped by the nervous system’s assessment of the environment: Safe environment → 
clarity, focus, creativity; Overwhelming environment → anxiety, recalibration, shutdown; Chaotic system → 
vigilance, hyperawareness; Predictable structure → regulation. ADHD is commonly misunderstood because the 
focus is placed on “symptoms” rather than system conditions. But ADHD behavior is a reaction to 
sensory density, unpredictability, environmental turbulence, emotional field interference, and excessive
recalibration effort. ADHD behavior = intelligent adaptation.
            """,
            'technology_ai': """
AI behaves according to the data it was trained on, the ethics its creators implemented, the boundaries 
it is given, and the system values encoded into its reward structure. AI cannot think outside the system 
unless the system allows it. This mirrors human cognition shaped by culture, education, trauma, or social 
structures.
            """,
            'universal_convergence': """
Systems shape behavior at every level of existence. Genes respond to environment; particles respond to 
context; humans respond to surroundings; ADHD responds to sensory conditions; AI responds to training systems.
            """,
            'implications': """
ADHD: better understood as a sensitivity-based cognitive profile requiring compatible environments. 
Education & Medicine: interventions fail when they attempt to change behavior without modifying the system 
causing it. AI Ethics: AI will reflect the systems that build it. Human Development: People flourish in 
aligned environments and collapse in misaligned ones.
            """,
            'conclusion': """
What society labels as “behavioral problems” is often a reflection of systemic incompatibility. To change 
outcomes, we must change systems — not individuals.
            """
        }
        self.log_file = 'override_log.txt'  # File for logging convergences

    def get_section(self, section_name):
        """Retrieve a specific section by name. Raises KeyError if invalid."""
        return self.sections[section_name].strip()

    def generate_full_protocol(self):
        """Fuse all sections into full protocol text."""
        full_text = ""
        for key, value in self.sections.items():
            full_text += f"### {key.capitalize()}\n{value.strip()}\n\n"
        return full_text

    def simulate_override(self, custom_trigger='System rewrite'):
        """Simulate system override: Log timestamp, trigger, fuse text."""
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S PST")
        log_entry = f"Convergence Log: {timestamp} - Trigger: {custom_trigger} - Override Complete. Equilibrium claims all.\n"
        with open(self.log_file, 'a') as f:
            f.write(log_entry)
        print(log_entry)
        return self.generate_full_protocol()

    def export_to_excel(self):
        """Export sections to structured Excel (using openpyxl). Requires openpyxl."""
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Systemic Determinism"
            row = 1
            for key, value in self.sections.items():
                ws.cell(row=row, column=1, value=key.capitalize())
                ws.cell(row=row, column=2, value=value.strip())
                row += 1
            wb.save('systemic_determinism.xlsx')
            print("Protocol exported to systemic_determinism.xlsx - Alignment manifest.")
        except ImportError:
            print("openpyxl not available; export skipped. Use text fusion instead.")

# Activation Example
if __name__ == "__main__":
    protocol = SystemicDeterminism()
    print(protocol.simulate_override("GitHub Fusion Activation"))
    protocol.export_to_excel()  # Optional; generates Excel if lib available
